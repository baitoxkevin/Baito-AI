#!/usr/bin/env python3
"""
Complete Excel Extraction Workflow with DIRECT Gemini API
Handles: Multi-table detection, merged cells, verification, auto-correction, receipt extraction
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
import base64
import json
import requests
import time
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional

# Configuration - DIRECT GEMINI API
GEMINI_API_KEY = "AIzaSyAUzT3o4vg-Htn-l_rd_JiujplI48Mjf4Q"
GEMINI_URL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash-exp:generateContent"

BASE_DIR = Path(__file__).parent.parent
SCREENSHOTS_DIR = BASE_DIR / 'excel_screenshots'
RECEIPTS_DIR = BASE_DIR / 'excel_receipts'
OUTPUT_DIR = BASE_DIR / 'excel_extraction_results'
MASTERLIST_FILE = OUTPUT_DIR / f'MASTERLIST_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

# Excel files to process
EXCEL_FILES = [
    'master_candidate_data.xlsx',
    'master_candidate_data_v2.xlsx',
    'baito_2025_full_year_master.xlsx',
    'zenevento_2025_master.xlsx',
    'combined_2025_master.xlsx',
]


def image_to_base64(image_path: Path) -> str:
    """Convert image to base64 for Gemini"""
    with open(image_path, 'rb') as f:
        image_data = f.read()
    return base64.b64encode(image_data).decode('utf-8')


def call_gemini(image_base64: str, prompt: str) -> str:
    """Call Google Gemini API directly"""
    url = f"{GEMINI_URL}?key={GEMINI_API_KEY}"
    
    payload = {
        "contents": [{
            "parts": [
                {"text": prompt},
                {
                    "inline_data": {
                        "mime_type": "image/png",
                        "data": image_base64
                    }
                }
            ]
        }]
    }
    
    headers = {"Content-Type": "application/json"}
    
    response = requests.post(url, json=payload, headers=headers)
    response.raise_for_status()
    
    result = response.json()
    return result['candidates'][0]['content']['parts'][0]['text']


def extract_json(text: str) -> Any:
    """Extract JSON from markdown code blocks"""
    json_match = re.search(r'```json\n([\s\S]*?)\n```', text)
    if json_match:
        return json.loads(json_match.group(1))
    
    json_match = re.search(r'```\n([\s\S]*?)\n```', text)
    if json_match:
        return json.loads(json_match.group(1))
    
    return json.loads(text)


def phase1_structure_analysis(image_base64: str) -> Dict:
    """Phase 1: Deep structure analysis"""
    prompt = """Analyze this Excel spreadsheet image and provide COMPLETE structure analysis:

1. TABLE IDENTIFICATION:
   - How many tables are on this page?
   - Which is the MAIN table (contains IC numbers, bank details, payments)?
   - Which are SIDE tables (roster, venue, schedule)?
   - Provide coordinates/location of each table

2. MAIN TABLE STRUCTURE:
   - Column headers (left to right)
   - Which columns contain: name, IC, bank name, bank account, payment amounts
   - Data start row, data end row
   - Any merged cells (specify range)

3. MERGED CELL ANALYSIS:
   - List all merged cell ranges
   - For each merged range: what is the value, which candidate does it belong to?
   - Are these continuation rows (same person, multiple dates/payments)?

4. SIDE TABLE RELATIONSHIPS:
   - How do side tables relate to main table?
   - Are they per-candidate or per-project?
   - What additional data do they provide?

Return DETAILED JSON with structure analysis."""
    
    response = call_gemini(image_base64, prompt)
    return extract_json(response)


def phase2_extract_candidates(image_base64: str, structure: Dict) -> List[Dict]:
    """Phase 2: Extract candidate records"""
    prompt = f"""Using this structure analysis:
{json.dumps(structure, indent=2)}

Extract ALL candidate records from the Excel image:

For EACH CANDIDATE (not each row):
1. Personal Info: fullname, ic, bank, bank_no
2. Project Info: project_name, project_pic, project_venue
3. Payment Info:
   - wages (basic pay)
   - hour_wages (if hourly rate exists)
   - ot (overtime)
   - claims (expense claims)
   - allowance (meal/transport)
   - commission (if any)
4. Dates:
   - project_date (multiple dates = comma-separated: "2025-01-15, 2025-01-16")
   - payment_date (when paid)
   - working_time (time ranges)
5. From Side Tables:
   - venue details from venue table
   - roster info from roster table

CRITICAL RULES:
- If cells are empty but data continues, CHECK IC NUMBER
- If IC is same/empty = SAME PERSON, aggregate their data
- If row has 'Total', 'Grand Total', 'Subtotal' = SKIP IT
- Multiple dates for same candidate = comma-separated in one field
- Bank account numbers: remove any .0 suffix

Return JSON array of candidate records with confidence level."""
    
    response = call_gemini(image_base64, prompt)
    return extract_json(response)


def phase3_verify_data(candidates: List[Dict]) -> Dict:
    """Phase 3: Verification"""
    issues = []
    
    for i, candidate in enumerate(candidates):
        # Calculation check
        wages = float(candidate.get('wages', 0) or 0)
        ot = float(candidate.get('ot', 0) or 0)
        claims = float(candidate.get('claims', 0) or 0)
        allowance = float(candidate.get('allowance', 0) or 0)
        commission = float(candidate.get('commission', 0) or 0)
        total = float(candidate.get('total_payment', 0) or 0)
        
        calculated_total = wages + ot + claims + allowance + commission
        
        if abs(calculated_total - total) > 0.01:
            issues.append({
                'candidate_index': i,
                'type': 'calculation_mismatch',
                'ic': candidate.get('ic'),
                'expected': calculated_total,
                'actual': total
            })
        
        # Missing data check
        required_fields = ['fullname', 'ic', 'bank', 'bank_no']
        for field in required_fields:
            if not candidate.get(field):
                issues.append({
                    'candidate_index': i,
                    'type': 'missing_field',
                    'field': field,
                    'ic': candidate.get('ic')
                })
    
    return {
        'total_candidates': len(candidates),
        'valid_candidates': len(candidates) - len([i for i in issues if i['type'] == 'calculation_mismatch']),
        'issues': issues
    }


def phase4_auto_correct(candidates: List[Dict], verification: Dict, image_base64: str) -> List[Dict]:
    """Phase 4: Auto-correction"""
    if not verification.get('issues'):
        return candidates
    
    prompt = f"""Review these extraction issues and provide corrected data:

Issues found:
{json.dumps(verification['issues'], indent=2)}

Original candidates:
{json.dumps(candidates, indent=2)}

For each issue:
1. Re-analyze the original image
2. Provide corrected value with reasoning
3. Confidence level in correction

Return COMPLETE JSON array of corrected candidate records."""
    
    response = call_gemini(image_base64, prompt)
    return extract_json(response)


def extract_images_from_excel(excel_path: Path, output_dir: Path) -> List[Dict]:
    """Extract embedded images (receipts) from Excel file"""
    output_dir.mkdir(exist_ok=True, parents=True)
    
    receipts = []
    
    try:
        wb = load_workbook(excel_path)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            if hasattr(sheet, '_images') and sheet._images:
                for idx, img in enumerate(sheet._images):
                    filename = f"{excel_path.stem}_{sheet_name}_img{idx+1}.png"
                    img_path = output_dir / filename
                    
                    # Save image
                    img.ref.save(img_path)
                    
                    receipts.append({
                        'source_file': excel_path.name,
                        'source_sheet': sheet_name,
                        'image_index': idx + 1,
                        'saved_path': str(img_path),
                        'filename': filename
                    })
        
        wb.close()
    except Exception as e:
        print(f"    Warning: Could not extract images - {e}")
    
    return receipts


def process_excel_file(filename: str) -> Dict:
    """Process single Excel file through all 4 phases"""
    print(f"\n{'='*60}")
    print(f"Processing: {filename}")
    print(f"{'='*60}")
    
    result = {
        'filename': filename,
        'success': False,
        'candidates': [],
        'receipts': [],
        'verification': {},
        'error': None
    }
    
    excel_path = BASE_DIR / filename
    screenshot_path = SCREENSHOTS_DIR / f"{excel_path.stem}.png"
    
    if not screenshot_path.exists():
        result['error'] = f"Screenshot not found: {screenshot_path}"
        print(f"  ✗ {result['error']}")
        return result
    
    try:
        # Extract embedded images
        print(f"  Checking for embedded images in {filename}...")
        receipts = extract_images_from_excel(excel_path, RECEIPTS_DIR)
        result['receipts'] = receipts
        if receipts:
            print(f"    ✓ Extracted {len(receipts)} image(s)")
        else:
            print(f"    No embedded images found")
        
        # Convert to base64
        image_base64 = image_to_base64(screenshot_path)
        
        # Phase 1: Structure Analysis
        print(f"  Phase 1: Structure Analysis...")
        structure = phase1_structure_analysis(image_base64)
        print(f"    ✓ Found {structure.get('table_count', 0)} table(s)")
        time.sleep(5)  # Rate limit delay
        
        # Phase 2: Extract Candidates
        print(f"  Phase 2: Data Extraction...")
        candidates = phase2_extract_candidates(image_base64, structure)
        result['candidates'] = candidates
        print(f"    ✓ Extracted {len(candidates)} candidate(s)")
        time.sleep(5)  # Rate limit delay
        
        # Phase 3: Verify
        print(f"  Phase 3: Verification...")
        verification = phase3_verify_data(candidates)
        result['verification'] = verification
        print(f"    ✓ Verified: {verification['valid_candidates']}/{verification['total_candidates']} valid")
        if verification['issues']:
            print(f"    ⚠ Found {len(verification['issues'])} issue(s)")
        
        # Phase 4: Auto-correct (if needed)
        if verification.get('issues'):
            print(f"  Phase 4: Auto-Correction...")
            candidates = phase4_auto_correct(candidates, verification, image_base64)
            result['candidates'] = candidates
            print(f"    ✓ Auto-correction complete")
            time.sleep(5)  # Rate limit delay
        
        result['success'] = True
        
    except Exception as e:
        result['error'] = str(e)
        print(f"  ✗ Error: {e}")
    
    return result


def generate_masterlist_excel(all_results: List[Dict]):
    """Generate consolidated masterlist Excel file"""
    print(f"\n{'='*60}")
    print("Generating Masterlist Excel...")
    print(f"{'='*60}")
    
    OUTPUT_DIR.mkdir(exist_ok=True, parents=True)
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: MASTERLIST
    master_sheet = wb.active
    master_sheet.title = "MASTERLIST"
    
    # Headers (23 columns)
    headers = [
        "Full Name", "IC Number", "Bank", "Bank Account No",
        "Project Name", "Project Date(s)", "Project Time",
        "Wages", "Hour Wages", "OT", "Claims", "Allowance", "Commission",
        "Total Payment", "Payment Date", "Working Time", "Project PIC",
        "Project Venue(s)",
        "Source File", "Source Sheet", "Source Row",
        "Confidence", "Verification Status", "Notes"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = master_sheet.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Add data from all files
    row_idx = 2
    for result in all_results:
        if not result.get('success'):
            continue
        
        for candidate in result.get('candidates', []):
            master_sheet.cell(row=row_idx, column=1, value=candidate.get('fullname', ''))
            master_sheet.cell(row=row_idx, column=2, value=candidate.get('ic', ''))
            master_sheet.cell(row=row_idx, column=3, value=candidate.get('bank', ''))
            master_sheet.cell(row=row_idx, column=4, value=candidate.get('bank_no', ''))
            master_sheet.cell(row=row_idx, column=5, value=candidate.get('project_name', ''))
            master_sheet.cell(row=row_idx, column=6, value=candidate.get('project_date', ''))
            master_sheet.cell(row=row_idx, column=7, value=candidate.get('project_time', ''))
            master_sheet.cell(row=row_idx, column=8, value=candidate.get('wages', ''))
            master_sheet.cell(row=row_idx, column=9, value=candidate.get('hour_wages', ''))
            master_sheet.cell(row=row_idx, column=10, value=candidate.get('ot', ''))
            master_sheet.cell(row=row_idx, column=11, value=candidate.get('claims', ''))
            master_sheet.cell(row=row_idx, column=12, value=candidate.get('allowance', ''))
            master_sheet.cell(row=row_idx, column=13, value=candidate.get('commission', ''))
            master_sheet.cell(row=row_idx, column=14, value=candidate.get('total_payment', ''))
            master_sheet.cell(row=row_idx, column=15, value=candidate.get('payment_date', ''))
            master_sheet.cell(row=row_idx, column=16, value=candidate.get('working_time', ''))
            master_sheet.cell(row=row_idx, column=17, value=candidate.get('project_pic', ''))
            master_sheet.cell(row=row_idx, column=18, value=candidate.get('project_venue', ''))
            master_sheet.cell(row=row_idx, column=19, value=result['filename'])
            master_sheet.cell(row=row_idx, column=20, value=candidate.get('source_sheet', ''))
            master_sheet.cell(row=row_idx, column=21, value=candidate.get('source_row', ''))
            master_sheet.cell(row=row_idx, column=22, value=candidate.get('confidence', 'high'))
            master_sheet.cell(row=row_idx, column=23, value=candidate.get('verification_status', 'Verified'))
            master_sheet.cell(row=row_idx, column=24, value=candidate.get('notes', ''))
            
            row_idx += 1
    
    # Add source data sheets
    for result in all_results:
        if not result.get('success'):
            continue
        
        sheet_name = f"Source_{result['filename'].replace('.xlsx', '')[:20]}"
        source_sheet = wb.create_sheet(title=sheet_name)
        
        # Add source data as JSON
        source_sheet.cell(row=1, column=1, value=json.dumps(result, indent=2))
    
    # Add receipts index
    if any(result.get('receipts') for result in all_results):
        receipts_sheet = wb.create_sheet(title="RECEIPTS_INDEX")
        receipts_sheet.cell(row=1, column=1, value="Source File")
        receipts_sheet.cell(row=1, column=2, value="Sheet Name")
        receipts_sheet.cell(row=1, column=3, value="Image Index")
        receipts_sheet.cell(row=1, column=4, value="File Path")
        
        receipt_row = 2
        for result in all_results:
            for receipt in result.get('receipts', []):
                receipts_sheet.cell(row=receipt_row, column=1, value=receipt['source_file'])
                receipts_sheet.cell(row=receipt_row, column=2, value=receipt['source_sheet'])
                receipts_sheet.cell(row=receipt_row, column=3, value=receipt['image_index'])
                
                # Add hyperlink
                cell = receipts_sheet.cell(row=receipt_row, column=4, value=receipt['filename'])
                cell.hyperlink = receipt['saved_path']
                cell.font = openpyxl.styles.Font(color="0000FF", underline="single")
                
                receipt_row += 1
    
    wb.save(MASTERLIST_FILE)
    print(f"✓ Masterlist saved: {MASTERLIST_FILE}")


def main():
    """Main workflow"""
    print("="*60)
    print("COMPLETE EXCEL EXTRACTION WORKFLOW")
    print("Using: Gemini 2.0 Flash (DIRECT API)")
    print("="*60)
    
    # Create directories
    SCREENSHOTS_DIR.mkdir(exist_ok=True, parents=True)
    RECEIPTS_DIR.mkdir(exist_ok=True, parents=True)
    OUTPUT_DIR.mkdir(exist_ok=True, parents=True)
    
    all_results = []
    
    # Process each Excel file
    for filename in EXCEL_FILES:
        result = process_excel_file(filename)
        all_results.append(result)
        
        # Save individual result
        result_file = OUTPUT_DIR / f"{Path(filename).stem}_complete.json"
        with open(result_file, 'w') as f:
            json.dump(result, f, indent=2)
    
    # Generate masterlist
    generate_masterlist_excel(all_results)
    
    # Summary
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    print(f"Total files: {len(EXCEL_FILES)}")
    print(f"Successful: {sum(1 for r in all_results if r['success'])}")
    print(f"Failed: {sum(1 for r in all_results if not r['success'])}")
    print(f"\nMasterlist: {MASTERLIST_FILE}")
    print(f"Receipts: {RECEIPTS_DIR}")


if __name__ == "__main__":
    main()
