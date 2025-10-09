#!/usr/bin/env python3
"""
Excel Vision Processor
Converts Excel files to images and sends them to n8n workflow for AI extraction
"""

import openpyxl
from PIL import Image, ImageDraw, ImageFont
import base64
import json
import requests
import time
from pathlib import Path
from datetime import datetime

# Configuration - USE TEST WEBHOOK (works regardless of activation)
WEBHOOK_URL = 'http://localhost:5678/webhook-test/vision-to-excel'
BASE_DIR = Path(__file__).parent.parent
OUTPUT_DIR = BASE_DIR / 'excel_extraction_results'
SCREENSHOTS_DIR = BASE_DIR / 'excel_screenshots'

# Excel files to process
EXCEL_FILES = [
    'master_candidate_data.xlsx',
    'master_candidate_data_v2.xlsx',
    'baito_2025_full_year_master.xlsx',
    'zenevento_2025_master.xlsx',
    'combined_2025_master.xlsx',
    'baito_2025_full_year_master_CORRECTED.xlsx',
    'baito_2025_VALIDATED_v3.xlsx',
    'baito_2025_VALIDATED_v3_FIXED.xlsx',
    'baito_2025_FIXED_v3.2.xlsx',
    'baito_2025_COMPLETE_v3.xlsx'
]


def excel_to_image(excel_path, output_path):
    """Convert Excel file to PNG image"""
    print(f"Converting {excel_path.name} to image...")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active

    # Calculate column widths
    col_widths = {}
    for col_idx in range(1, sheet.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        width = sheet.column_dimensions[col_letter].width
        col_widths[col_idx] = int((width if width else 10) * 7)

    # Calculate row heights
    row_heights = {}
    for row_idx in range(1, sheet.max_row + 1):
        height = sheet.row_dimensions[row_idx].height
        row_heights[row_idx] = int((height if height else 15) * 1.3)

    # Calculate image dimensions
    img_width = min(sum(col_widths.values()) + 20, 4000)  # Max width
    img_height = min(sum(row_heights.values()) + 20, 4000)  # Max height

    # Create image
    img = Image.new('RGB', (img_width, img_height), 'white')
    draw = ImageDraw.Draw(img)

    # Try to load a nice font
    try:
        font = ImageFont.truetype('/System/Library/Fonts/Helvetica.ttc', 11)
        font_bold = ImageFont.truetype('/System/Library/Fonts/Helvetica.ttc', 12)
    except:
        try:
            font = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', 11)
            font_bold = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf', 12)
        except:
            font = ImageFont.load_default()
            font_bold = font

    # Draw cells
    y_offset = 10
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 100)), 1):
        if y_offset > img_height - 20:
            break

        x_offset = 10
        for col_idx, cell in enumerate(row, 1):
            if x_offset > img_width - 20:
                break

            cell_width = col_widths.get(col_idx, 70)
            cell_height = row_heights.get(row_idx, 20)

            # Draw cell border
            draw.rectangle(
                [x_offset, y_offset, x_offset + cell_width, y_offset + cell_height],
                outline='#CCCCCC',
                width=1
            )

            # Highlight header row
            if row_idx == 1:
                draw.rectangle(
                    [x_offset, y_offset, x_offset + cell_width, y_offset + cell_height],
                    fill='#F0F0F0',
                    outline='#999999',
                    width=1
                )

            # Draw cell value
            if cell.value is not None:
                text = str(cell.value)[:40]  # Truncate long text
                cell_font = font_bold if row_idx == 1 else font

                # Text wrapping for long content
                if len(text) > 20:
                    words = text.split()
                    lines = []
                    current_line = []
                    for word in words:
                        test_line = ' '.join(current_line + [word])
                        if len(test_line) <= 20:
                            current_line.append(word)
                        else:
                            if current_line:
                                lines.append(' '.join(current_line))
                            current_line = [word]
                    if current_line:
                        lines.append(' '.join(current_line))

                    for i, line in enumerate(lines[:2]):  # Max 2 lines
                        draw.text(
                            (x_offset + 3, y_offset + 3 + i * 12),
                            line,
                            fill='black',
                            font=cell_font
                        )
                else:
                    draw.text(
                        (x_offset + 3, y_offset + 3),
                        text,
                        fill='black',
                        font=cell_font
                    )

            x_offset += cell_width
        y_offset += cell_height

    img.save(output_path, 'PNG', optimize=True)
    print(f"✓ Image created: {output_path.name}")
    return output_path


def image_to_data_url(image_path):
    """Convert image to base64 data URL"""
    with open(image_path, 'rb') as f:
        image_data = f.read()
    base64_data = base64.b64encode(image_data).decode('utf-8')
    return f"data:image/png;base64,{base64_data}"


def send_to_workflow(image_data_url, metadata):
    """Send image to n8n webhook for processing"""
    print(f"Sending {metadata['filename']} to n8n workflow...")

    payload = {
        'image': image_data_url,
        'metadata': metadata
    }

    response = requests.post(
        WEBHOOK_URL,
        json=payload,
        headers={'Content-Type': 'application/json'},
        timeout=180  # 3 minutes timeout for AI processing
    )

    if response.status_code == 200:
        try:
            return response.json()
        except:
            # Response might be empty, return success indicator
            return {"success": True, "message": "Workflow executed"}
    else:
        response.raise_for_status()


def process_excel_file(filename):
    """Process a single Excel file"""
    excel_path = BASE_DIR / filename
    image_path = SCREENSHOTS_DIR / f"{excel_path.stem}.png"
    result_path = OUTPUT_DIR / f"{excel_path.stem}_result.json"

    print(f"\n=== Processing: {filename} ===")

    try:
        # Check if file exists
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        # Step 1: Convert Excel to image
        excel_to_image(excel_path, image_path)

        # Step 2: Convert to data URL
        image_data_url = image_to_data_url(image_path)

        # Step 3: Send to n8n workflow
        result = send_to_workflow(image_data_url, {
            'filename': filename,
            'processedAt': datetime.now().isoformat(),
            'imagePath': str(image_path)
        })

        # Step 4: Save result
        with open(result_path, 'w') as f:
            json.dump(result, f, indent=2)

        print(f"✓ Extracted {result.get('extractedRecords', 0)} records")
        print(f"✓ Result saved to: {result_path.name}")

        return {
            'success': True,
            'filename': filename,
            'records': result.get('extractedRecords', 0),
            'validation': result.get('validationReport', {})
        }

    except Exception as error:
        print(f"✗ Error processing {filename}: {error}")
        return {
            'success': False,
            'filename': filename,
            'error': str(error)
        }


def main():
    """Main execution"""
    print('Excel Vision Processor')
    print('=' * 50 + '\n')

    # Create output directories
    OUTPUT_DIR.mkdir(exist_ok=True)
    SCREENSHOTS_DIR.mkdir(exist_ok=True)

    # Check if webhook is accessible
    print('Checking n8n webhook...')
    try:
        response = requests.get(WEBHOOK_URL, timeout=5)
        print(f"✓ Webhook accessible (status: {response.status_code})\n")
    except Exception as error:
        print('✗ Cannot connect to n8n webhook!')
        print('  Make sure:')
        print('  1. n8n is running at http://localhost:5678')
        print('  2. Workflow "Excel Vision Extractor" is ACTIVE')
        print('  3. Webhook path is "vision-to-excel"\n')
        print(f'  Error: {error}\n')
        return

    # Process all Excel files
    results = []
    for filename in EXCEL_FILES:
        result = process_excel_file(filename)
        results.append(result)

        # Add delay between requests to avoid rate limiting
        time.sleep(3)

    # Summary
    print('\n' + '=' * 50)
    print('=== Summary ===')
    successful = [r for r in results if r['success']]
    failed = [r for r in results if not r['success']]

    print(f"Total files: {len(results)}")
    print(f"Successful: {len(successful)}")
    print(f"Failed: {len(failed)}")

    if successful:
        total_records = sum(r.get('records', 0) for r in successful)
        print(f"Total records extracted: {total_records}")

    if failed:
        print('\nFailed files:')
        for r in failed:
            print(f"  - {r['filename']}: {r['error']}")

    print(f"\nResults saved in: {OUTPUT_DIR}")
    print(f"Screenshots saved in: {SCREENSHOTS_DIR}")


if __name__ == '__main__':
    main()
