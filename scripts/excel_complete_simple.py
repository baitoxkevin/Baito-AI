#!/usr/bin/env python3
"""
Simple Complete Excel Extraction with Gemini Direct API
"""
import requests
import base64
import json
import re
import openpyxl
from pathlib import Path
from datetime import datetime

GEMINI_API_KEY = "AIzaSyAUzT3o4vg-Htn-l_rd_JiujplI48Mjf4Q"
BASE_DIR = Path("/Users/baito.kevin/Downloads/dev/BMAD-METHOD/Baito/Baito-AI")
SCREENSHOTS_DIR = BASE_DIR / "excel_screenshots"
OUTPUT_DIR = BASE_DIR / "excel_extraction_results"

EXCEL_FILES = [
    'master_candidate_data.xlsx',
    'master_candidate_data_v2.xlsx',
    'baito_2025_full_year_master.xlsx',
    'zenevento_2025_master.xlsx',
    'combined_2025_master.xlsx',
]

def call_gemini(image_base64, prompt):
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash-exp:generateContent?key={GEMINI_API_KEY}"
    payload = {
        "contents": [{
            "parts": [
                {"text": prompt},
                {"inline_data": {"mime_type": "image/png", "data": image_base64}}
            ]
        }]
    }
    response = requests.post(url, json=payload, headers={"Content-Type": "application/json"}, timeout=180)
    response.raise_for_status()
    return response.json()['candidates'][0]['content']['parts'][0]['text']

def extract_json(text):
    json_match = re.search(r'```json\n([\s\S]*?)\n```', text)
    if json_match:
        return json.loads(json_match.group(1))
    json_match = re.search(r'```\n([\s\S]*?)\n```', text)
    if json_match:
        return json.loads(json_match.group(1))
    return json.loads(text)

def process_file(filename):
    print(f"\n{'='*60}")
    print(f"Processing: {filename}")
    print(f"{'='*60}")
    
    screenshot = SCREENSHOTS_DIR / f"{Path(filename).stem}.png"
    if not screenshot.exists():
        print(f"  ✗ Screenshot not found")
        return None
    
    print(f"  Loading image...")
    with open(screenshot, 'rb') as f:
        image_base64 = base64.b64encode(f.read()).decode('utf-8')
    
    prompt = """Extract ALL candidates from this Excel payroll sheet.

For EACH candidate, extract these fields:
- fullname
- ic (IC number)
- bank
- bank_no (bank account number - remove .0 suffix if present)
- project_name
- project_date (if multiple dates for same person, use comma-separated: "2025-01-15, 2025-01-16")
- project_time
- wages
- hour_wages (if hourly rate column exists)
- ot (overtime amount)
- claims (expense claims)
- allowance (meal/transport allowance)
- commission (if any)
- total_payment (calculate: wages + ot + claims + allowance + commission)
- payment_date
- working_time
- project_pic (project person in charge)
- project_venue (if multiple venues, comma-separated)

IMPORTANT RULES:
- If a person appears in multiple rows (same IC or name), combine their data into ONE record with multiple dates
- Skip rows with "Total", "Grand Total", "Subtotal"
- If a field doesn't exist in the sheet, use null or empty string

Return ONLY a JSON array of candidate objects."""

    print(f"  Extracting candidates...")
    try:
        response_text = call_gemini(image_base64, prompt)
        candidates = extract_json(response_text)
        print(f"  ✓ Extracted {len(candidates)} candidates")
        
        # Save individual result
        output_file = OUTPUT_DIR / f"{Path(filename).stem}_result.json"
        with open(output_file, 'w') as f:
            json.dump(candidates, f, indent=2)
        
        return {
            'filename': filename,
            'success': True,
            'candidates': candidates
        }
    except Exception as e:
        print(f"  ✗ Error: {e}")
        return {
            'filename': filename,
            'success': False,
            'error': str(e),
            'candidates': []
        }

def generate_masterlist(all_results):
    print(f"\n{'='*60}")
    print("Generating Masterlist Excel...")
    print(f"{'='*60}")
    
    OUTPUT_DIR.mkdir(exist_ok=True, parents=True)
    masterlist_file = OUTPUT_DIR / f"MASTERLIST_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MASTERLIST"
    
    # Headers
    headers = [
        "Full Name", "IC Number", "Bank", "Bank Account No",
        "Project Name", "Project Date(s)", "Project Time",
        "Wages", "Hour Wages", "OT", "Claims", "Allowance", "Commission",
        "Total Payment", "Payment Date", "Working Time", "Project PIC",
        "Project Venue(s)", "Source File"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
    
    row_idx = 2
    for result in all_results:
        if not result['success']:
            continue
        
        for candidate in result['candidates']:
            ws.cell(row=row_idx, column=1, value=candidate.get('fullname', ''))
            ws.cell(row=row_idx, column=2, value=candidate.get('ic', ''))
            ws.cell(row=row_idx, column=3, value=candidate.get('bank', ''))
            ws.cell(row=row_idx, column=4, value=candidate.get('bank_no', ''))
            ws.cell(row=row_idx, column=5, value=candidate.get('project_name', ''))
            ws.cell(row=row_idx, column=6, value=candidate.get('project_date', ''))
            ws.cell(row=row_idx, column=7, value=candidate.get('project_time', ''))
            ws.cell(row=row_idx, column=8, value=candidate.get('wages', ''))
            ws.cell(row=row_idx, column=9, value=candidate.get('hour_wages', ''))
            ws.cell(row=row_idx, column=10, value=candidate.get('ot', ''))
            ws.cell(row=row_idx, column=11, value=candidate.get('claims', ''))
            ws.cell(row=row_idx, column=12, value=candidate.get('allowance', ''))
            ws.cell(row=row_idx, column=13, value=candidate.get('commission', ''))
            ws.cell(row=row_idx, column=14, value=candidate.get('total_payment', ''))
            ws.cell(row=row_idx, column=15, value=candidate.get('payment_date', ''))
            ws.cell(row=row_idx, column=16, value=candidate.get('working_time', ''))
            ws.cell(row=row_idx, column=17, value=candidate.get('project_pic', ''))
            ws.cell(row=row_idx, column=18, value=candidate.get('project_venue', ''))
            ws.cell(row=row_idx, column=19, value=result['filename'])
            row_idx += 1
    
    wb.save(masterlist_file)
    print(f"✓ Masterlist saved: {masterlist_file}")
    return masterlist_file

def main():
    print("="*60)
    print("COMPLETE EXCEL EXTRACTION - SIMPLE & EFFECTIVE")
    print("Using: Gemini 2.0 Flash (Direct API)")
    print("="*60)
    
    all_results = []
    
    for filename in EXCEL_FILES:
        result = process_file(filename)
        if result:
            all_results.append(result)
        print("  Waiting 10 seconds before next file...")
        import time
        time.sleep(10)
    
    masterlist_file = generate_masterlist(all_results)
    
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    print(f"Total files: {len(EXCEL_FILES)}")
    print(f"Successful: {sum(1 for r in all_results if r['success'])}")
    print(f"Failed: {sum(1 for r in all_results if not r['success'])}")
    print(f"Total candidates: {sum(len(r['candidates']) for r in all_results)}")
    print(f"\nMasterlist: {masterlist_file}")

if __name__ == "__main__":
    main()
