#!/usr/bin/env python3
"""
Complete Excel Extraction Workflow with Gemini Vision
Handles: Multi-table detection, merged cells, verification, auto-correction, receipt extraction
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
import base64
import json
import requests
import time
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional

# Configuration
OPENROUTER_API_KEY = "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"
OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"
MODEL = "google/gemini-2.0-flash-exp:free"

BASE_DIR = Path(__file__).parent.parent
SCREENSHOTS_DIR = BASE_DIR / 'excel_screenshots'
RECEIPTS_DIR = BASE_DIR / 'excel_receipts'
OUTPUT_DIR = BASE_DIR / 'excel_extraction_results'
MASTERLIST_FILE = OUTPUT_DIR / f'MASTERLIST_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

# Excel files to process
EXCEL_FILES = [
    'master_candidate_data.xlsx',
    'master_candidate_data_v2.xlsx',
    'baito_2025_full_year_master.xlsx',
    'zenevento_2025_master.xlsx',
    'combined_2025_master.xlsx',
]


def image_to_data_url(image_path: Path) -> str:
    """Convert image to base64 data URL"""
    with open(image_path, 'rb') as f:
        image_data = f.read()
    base64_data = base64.b64encode(image_data).decode('utf-8')
    return f"data:image/png;base64,{base64_data}"


def call_gemini(image_data_url: str, prompt: str) -> str:
    """Call Gemini via OpenRouter with vision"""
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "HTTP-Referer": "https://baito-ai.app",
        "X-Title": "Baito Excel Extractor",
        "Content-Type": "application/json"
    }

    payload = {
        "model": MODEL,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": image_data_url}}
            ]
        }]
    }

    response = requests.post(OPENROUTER_URL, headers=headers, json=payload, timeout=180)
    response.raise_for_status()
    result = response.json()
    return result['choices'][0]['message']['content']


def extract_json(text: str) -> Any:
    """Extract JSON from AI response (handles markdown)"""
    # Try markdown code blocks
    json_match = re.search(r'```json\n([\s\S]*?)\n```', text)
    if json_match:
        return json.loads(json_match.group(1))

    json_match = re.search(r'```\n([\s\S]*?)\n```', text)
    if json_match:
        return json.loads(json_match.group(1))

    # Try parsing whole response
    return json.loads(text)


def extract_images_from_excel(excel_path: Path, output_dir: Path) -> List[Dict]:
    """Extract embedded images (receipts) from Excel file"""
    print(f"  Checking for embedded images in {excel_path.name}...")

    extracted_images = []
    try:
        wb = load_workbook(excel_path)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Check if sheet has images
            if hasattr(sheet, '_images') and sheet._images:
                for idx, img in enumerate(sheet._images):
                    # Generate filename
                    img_filename = f"{excel_path.stem}_{sheet_name}_img{idx+1}.png"
                    img_path = output_dir / img_filename

                    # Save image
                    with open(img_path, 'wb') as f:
                        f.write(img.ref.getvalue())

                    extracted_images.append({
                        'filename': img_filename,
                        'path': str(img_path),
                        'sheet': sheet_name,
                        'index': idx,
                        'anchor': str(img.anchor) if hasattr(img, 'anchor') else None
                    })

                    print(f"    ✓ Extracted: {img_filename}")

        if not extracted_images:
            print(f"    No embedded images found")

    except Exception as e:
        print(f"    Error extracting images: {e}")

    return extracted_images


def phase1_structure_analysis(image_data_url: str) -> Dict:
    """Phase 1: Deep structure analysis"""
    print("  Phase 1: Structure Analysis...")

    prompt = """Analyze this Excel spreadsheet image and provide COMPLETE structure analysis:

1. TABLE IDENTIFICATION:
   - How many tables are on this page?
   - Which is the MAIN table (contains IC numbers, bank details, payments)?
   - Which are SIDE tables (roster, venue, schedule, notes)?
   - Provide approximate location of each table (top/middle/bottom, left/right)

2. MAIN TABLE STRUCTURE:
   - List ALL column headers (left to right)
   - Identify columns: name, IC, bank name, bank account, wages, OT, allowance, claims, commission, payment amounts, dates
   - Data starts at row number?
   - Data ends at row number?
   - Any merged cells? Which ranges?

3. MERGED CELL ANALYSIS:
   - List all merged/empty cell ranges that are continuation rows
   - For each: which candidate does it belong to? How do you know?

4. SIDE TABLE DETAILS:
   - What information do side tables contain?
   - How do they relate to the main table? (by date? by name? by project?)

Return ONLY valid JSON with this structure:
{
  "tableCount": number,
  "mainTable": {
    "location": "string",
    "headerRow": number,
    "dataStartRow": number,
    "dataEndRow": number,
    "columns": [{"index": 0, "name": "Name", "type": "text"}, ...],
    "mergedCells": [{"range": "A2:A5", "belongsTo": "candidate name/IC", "reason": "..."}]
  },
  "sideTables": [
    {
      "name": "Roster" or "Venue" or "Schedule",
      "location": "string",
      "purpose": "string",
      "linkMethod": "how it relates to main table"
    }
  ]
}"""

    response = call_gemini(image_data_url, prompt)
    return extract_json(response)


def phase2_extract_candidates(image_data_url: str, structure: Dict) -> List[Dict]:
    """Phase 2: Extract candidate records"""
    print("  Phase 2: Data Extraction...")

    prompt = f"""Using this structure analysis:
{json.dumps(structure, indent=2)}

Extract ALL candidate records from the Excel image.

CRITICAL RULES:
1. One record per CANDIDATE (not per row)
2. If cells are empty but data continues, check if IC is same/empty = SAME PERSON
3. For merged/continuation rows: aggregate dates, sum amounts if they're additions
4. Multiple project dates = comma-separated in one field
5. Skip summary rows (Total, Grand Total, Subtotal)
6. From side tables: extract venue, roster, schedule info and link to candidates

For EACH CANDIDATE extract:
{{
  "fullname": "string",
  "ic": "string (identity card number)",
  "bank": "string (bank name)",
  "bank_no": "string (account number, clean format)",
  "project_name": "string",
  "project_date": "comma-separated dates if multiple: 2025-01-15, 2025-01-16",
  "project_time": "string (working hours)",
  "wages": number,
  "hour_wages": number or null,
  "ot": number (overtime),
  "claims": number,
  "allowance": number,
  "commission": number,
  "payment_date": "YYYY-MM-DD",
  "working_time": "string (time range)",
  "project_pic": "string (person in charge)",
  "project_venue": "comma-separated if multiple venues",
  "source_row": "row number in original Excel",
  "confidence": "high/medium/low",
  "notes": "any special observations"
}}

Return JSON array of candidate records:
[{{ candidate1 }}, {{ candidate2 }}, ...]"""

    response = call_gemini(image_data_url, prompt)
    return extract_json(response)


def phase3_verify_data(candidates: List[Dict]) -> Dict:
    """Phase 3: Verification"""
    print("  Phase 3: Verification...")

    prompt = f"""Review this extracted candidate data and verify:

{json.dumps(candidates, indent=2)}

Check:
1. CALCULATIONS: wages + ot + claims + allowance + commission = sum? Flag mismatches.
2. DUPLICATES: Same IC appearing multiple times with different data?
3. DATA QUALITY:
   - Valid IC format?
   - Clean account numbers (no .0 suffix)?
   - Valid date formats?
   - Any obvious errors (Name='Total', Days=500)?
4. MISSING DATA: Which fields are empty? Is this expected or error?

Return JSON:
{{
  "totalRecords": number,
  "validRecords": number,
  "issues": [
    {{
      "candidate": "name or IC",
      "issue": "description",
      "severity": "critical/warning/info",
      "suggestedFix": "what should be corrected"
    }}
  ],
  "calculationErrors": [...],
  "duplicates": [...],
  "missingData": [...]
}}"""

    # Call Gemini without image (just analyzing JSON)
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "HTTP-Referer": "https://baito-ai.app",
        "Content-Type": "application/json"
    }

    payload = {
        "model": MODEL,
        "messages": [{
            "role": "user",
            "content": prompt
        }]
    }

    response = requests.post(OPENROUTER_URL, headers=headers, json=payload, timeout=120)
    response.raise_for_status()
    result = response.json()
    return extract_json(result['choices'][0]['message']['content'])


def phase4_auto_correct(candidates: List[Dict], verification: Dict, image_data_url: str) -> List[Dict]:
    """Phase 4: Auto-correction"""
    print("  Phase 4: Auto-Correction...")

    if not verification.get('issues'):
        print("    No issues to correct!")
        return candidates

    prompt = f"""Based on these verification issues:
{json.dumps(verification['issues'], indent=2)}

And the original extracted data:
{json.dumps(candidates, indent=2)}

CORRECT the data by re-analyzing this Excel image. For each issue:
1. Explain the problem
2. Provide corrected value
3. Reasoning for correction
4. Confidence level

Return corrected candidate array with same format, plus corrections field:
[
  {{
    ...all candidate fields...,
    "corrections": [
      {{
        "field": "field_name",
        "oldValue": "...",
        "newValue": "...",
        "reasoning": "...",
        "confidence": "high/medium/low"
      }}
    ]
  }}
]"""

    response = call_gemini(image_data_url, prompt)
    return extract_json(response)


def process_excel_file(filename: str) -> Dict:
    """Process single Excel file through all 4 phases"""
    excel_path = BASE_DIR / filename
    image_path = SCREENSHOTS_DIR / f"{Path(filename).stem}.png"

    print(f"\n{'='*60}")
    print(f"Processing: {filename}")
    print(f"{'='*60}")

    if not image_path.exists():
        return {
            "success": False,
            "error": "Screenshot not found",
            "filename": filename
        }

    try:
        # Extract embedded images (receipts)
        receipts = extract_images_from_excel(excel_path, RECEIPTS_DIR)

        # Convert screenshot to data URL
        image_data_url = image_to_data_url(image_path)

        # Phase 1: Structure Analysis
        structure = phase1_structure_analysis(image_data_url)
        print(f"    ✓ Found {structure.get('tableCount', 0)} table(s)")

        time.sleep(5)  # Rate limit delay

        # Phase 2: Extract Candidates
        candidates = phase2_extract_candidates(image_data_url, structure)
        print(f"    ✓ Extracted {len(candidates)} candidate(s)")

        time.sleep(5)

        # Phase 3: Verify
        verification = phase3_verify_data(candidates)
        print(f"    ✓ Verified: {verification.get('validRecords', 0)}/{verification.get('totalRecords', 0)} valid")

        if verification.get('issues'):
            print(f"    ⚠ Found {len(verification['issues'])} issue(s)")

            time.sleep(5)

            # Phase 4: Auto-correct
            candidates = phase4_auto_correct(candidates, verification, image_data_url)
            print(f"    ✓ Auto-correction complete")

        # Add source tracking
        for candidate in candidates:
            candidate['source_file'] = filename
            candidate['source_sheet'] = 'Sheet1'  # Update if multi-sheet
            candidate['extraction_date'] = datetime.now().isoformat()

        return {
            "success": True,
            "filename": filename,
            "structure": structure,
            "candidates": candidates,
            "verification": verification,
            "receipts": receipts,
            "recordCount": len(candidates)
        }

    except Exception as error:
        print(f"  ✗ Error: {error}")
        return {
            "success": False,
            "error": str(error),
            "filename": filename
        }


def generate_masterlist_excel(all_results: List[Dict]):
    """Generate consolidated masterlist Excel file"""
    print(f"\n{'='*60}")
    print("Generating Masterlist Excel...")
    print(f"{'='*60}")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Sheet 1: MASTERLIST (consolidated)
    ws_master = wb.active
    ws_master.title = "MASTERLIST"

    # Headers
    headers = [
        "Full Name", "IC Number", "Bank", "Bank Account No",
        "Project Name", "Project Date(s)", "Project Time",
        "Wages", "Hour Wages", "OT", "Claims", "Allowance", "Commission",
        "Total Payment", "Payment Date", "Working Time",
        "Project PIC", "Project Venue(s)",
        "Source File", "Source Sheet", "Source Row",
        "Confidence", "Verification Status", "Notes"
    ]

    # Write headers with styling
    for col_idx, header in enumerate(headers, 1):
        cell = ws_master.cell(1, col_idx, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    row_idx = 2
    for result in all_results:
        if not result.get('success'):
            continue

        for candidate in result.get('candidates', []):
            # Calculate total payment
            total = sum([
                candidate.get('wages', 0) or 0,
                candidate.get('ot', 0) or 0,
                candidate.get('claims', 0) or 0,
                candidate.get('allowance', 0) or 0,
                candidate.get('commission', 0) or 0
            ])

            row_data = [
                candidate.get('fullname'),
                candidate.get('ic'),
                candidate.get('bank'),
                candidate.get('bank_no'),
                candidate.get('project_name'),
                candidate.get('project_date'),
                candidate.get('project_time'),
                candidate.get('wages'),
                candidate.get('hour_wages'),
                candidate.get('ot'),
                candidate.get('claims'),
                candidate.get('allowance'),
                candidate.get('commission'),
                total,
                candidate.get('payment_date'),
                candidate.get('working_time'),
                candidate.get('project_pic'),
                candidate.get('project_venue'),
                candidate.get('source_file'),
                candidate.get('source_sheet'),
                candidate.get('source_row'),
                candidate.get('confidence'),
                result.get('verification', {}).get('validRecords', 0) > 0 and 'Verified' or 'Needs Review',
                candidate.get('notes')
            ]

            for col_idx, value in enumerate(row_data, 1):
                ws_master.cell(row_idx, col_idx, value)

            row_idx += 1

    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        ws_master.column_dimensions[get_column_letter(col_idx)].width = 15

    # Sheet 2+: Source data from each file
    for result in all_results:
        if not result.get('success'):
            continue

        sheet_name = Path(result['filename']).stem[:31]  # Excel sheet name limit
        ws_source = wb.create_sheet(f"Source_{sheet_name}")

        # Write raw extraction data
        ws_source.cell(1, 1, "RAW EXTRACTION DATA")
        ws_source.cell(2, 1, f"File: {result['filename']}")
        ws_source.cell(3, 1, f"Records: {result['recordCount']}")
        ws_source.cell(5, 1, json.dumps(result['candidates'], indent=2))

    # Last Sheet: Receipts Index
    if any(r.get('receipts') for r in all_results):
        ws_receipts = wb.create_sheet("RECEIPTS_INDEX")
        ws_receipts.cell(1, 1, "Receipt Filename")
        ws_receipts.cell(1, 2, "Source File")
        ws_receipts.cell(1, 3, "Sheet")
        ws_receipts.cell(1, 4, "File Path")

        receipt_row = 2
        for result in all_results:
            for receipt in result.get('receipts', []):
                ws_receipts.cell(receipt_row, 1, receipt['filename'])
                ws_receipts.cell(receipt_row, 2, result['filename'])
                ws_receipts.cell(receipt_row, 3, receipt['sheet'])
                ws_receipts.cell(receipt_row, 4, receipt['path'])
                receipt_row += 1

    # Save
    wb.save(MASTERLIST_FILE)
    print(f"✓ Masterlist saved: {MASTERLIST_FILE}")


def main():
    """Main execution"""
    print("="*60)
    print("COMPLETE EXCEL EXTRACTION WORKFLOW")
    print("Using: Gemini 2.0 Flash via OpenRouter")
    print("="*60)

    # Create directories
    SCREENSHOTS_DIR.mkdir(exist_ok=True)
    RECEIPTS_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)

    # Process all files
    all_results = []
    for filename in EXCEL_FILES:
        result = process_excel_file(filename)
        all_results.append(result)

        # Save individual result
        result_file = OUTPUT_DIR / f"{Path(filename).stem}_complete.json"
        with open(result_file, 'w') as f:
            json.dump(result, f, indent=2)

        time.sleep(15)  # Rate limit delay between files

    # Generate masterlist
    generate_masterlist_excel(all_results)

    # Summary
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    successful = [r for r in all_results if r.get('success')]
    failed = [r for r in all_results if not r.get('success')]

    print(f"Total files: {len(all_results)}")
    print(f"Successful: {len(successful)}")
    print(f"Failed: {len(failed)}")

    if successful:
        total_records = sum(r.get('recordCount', 0) for r in successful)
        total_receipts = sum(len(r.get('receipts', [])) for r in successful)
        print(f"Total candidate records: {total_records}")
        print(f"Total receipts extracted: {total_receipts}")

    print(f"\nMasterlist: {MASTERLIST_FILE}")
    print(f"Receipts: {RECEIPTS_DIR}")


if __name__ == '__main__':
    main()
