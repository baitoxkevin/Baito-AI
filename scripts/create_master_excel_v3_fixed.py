#!/usr/bin/env python3
"""
Complete Master Excel Creator - V3.2 FIXED
Fixes continuation row aggregation and column mapping issues
"""

import os
import glob
import re
from pathlib import Path
from datetime import datetime
import pandas as pd
import numpy as np
from collections import defaultdict
import openpyxl


def clean_ic_number(ic):
    """Clean and normalize IC number."""
    if pd.isna(ic):
        return None
    ic_str = str(ic).strip().split('\n')[0].split('(')[0].strip()
    ic_clean = ic_str.replace(' ', '').replace('-', '')
    return ic_clean if ic_clean and len(ic_clean) >= 6 else None


def clean_name(name):
    """Clean and normalize name."""
    if pd.isna(name):
        return None
    name_str = str(name).strip()
    name_str = ' '.join(name_str.split())
    name_str = re.sub(r'\([^)]*\)', '', name_str).strip()
    name_str = ' '.join(word.capitalize() for word in name_str.split())
    return name_str if name_str else None


def extract_alternate_name(name):
    """Extract alternate name from parentheses."""
    if pd.isna(name):
        return None
    name_str = str(name)
    match = re.search(r'\(([^)]+)\)', name_str)
    if match:
        alt = match.group(1).strip()
        return ' '.join(word.capitalize() for word in alt.split())
    return None


def clean_bank_name(bank):
    """Clean and normalize bank name."""
    if pd.isna(bank):
        return None
    return str(bank).strip() if str(bank).strip() else None


def clean_account(account):
    """Clean and normalize account number - NO .0"""
    if pd.isna(account):
        return None

    account_str = str(account).strip()

    # Handle scientific notation
    try:
        if 'e' in account_str.lower() or '.' in account_str:
            account_num = float(account_str)
            account_str = str(int(account_num))
    except:
        pass

    # Remove spaces, hyphens, .0
    account_str = account_str.replace(' ', '').replace('-', '')
    account_str = account_str.replace('.0', '').replace('.', '')

    return account_str if account_str and account_str.isdigit() else None


def safe_float(value):
    """Safely convert to float."""
    if pd.isna(value):
        return 0.0
    try:
        return float(value)
    except:
        return 0.0


def extract_project_metadata(excel_path, sheet_name):
    """Extract project-level metadata from sheet."""
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, nrows=10)

        metadata = {
            'project_date_range': None,
            'payment_due_date': None,
            'location': None,
            'time_schedule': None
        }

        for idx in range(min(10, len(df))):
            row = df.iloc[idx]
            row_str = ' '.join([str(val) for val in row if pd.notna(val)])
            row_str_lower = row_str.lower()

            # Extract date range
            if 'date:' in row_str_lower and not metadata['project_date_range']:
                date_match = re.search(r'date:\\s*(.+?)(?:payment|time|$)', row_str, re.IGNORECASE)
                if date_match:
                    metadata['project_date_range'] = date_match.group(1).strip().strip(',')

            # Extract payment due date
            if 'payment' in row_str_lower and not metadata['payment_due_date']:
                payment_match = re.search(r'payment\\s+(?:by\\s+)?(.+?)(?:\\s{2,}|$)', row_str, re.IGNORECASE)
                if payment_match:
                    metadata['payment_due_date'] = payment_match.group(1).strip().strip(',')

            # Extract location
            if 'location:' in row_str_lower and not metadata['location']:
                loc_match = re.search(r'location:\\s*(.+?)(?:\\s{2,}|$)', row_str, re.IGNORECASE)
                if loc_match:
                    metadata['location'] = loc_match.group(1).strip().strip(',')

            # Extract time
            if 'time:' in row_str_lower and not metadata['time_schedule']:
                time_match = re.search(r'time:\\s*(.+?)(?:payment|$)', row_str, re.IGNORECASE)
                if time_match:
                    metadata['time_schedule'] = time_match.group(1).strip().strip(',')

        return metadata
    except:
        return {
            'project_date_range': None,
            'payment_due_date': None,
            'location': None,
            'time_schedule': None
        }


def extract_month_from_path(path):
    """Extract month from file path."""
    path_str = str(path).lower()
    months = ['jan', 'feb', 'march', 'april', 'may', 'june', 'july', 'aug', 'sep', 'oct', 'nov', 'dec']
    for month in months:
        if month in path_str:
            return month.capitalize()
    return 'Unknown'


def identify_columns(df):
    """
    Identify column purposes with better mapping.
    Returns dict of column purposes.
    """
    cols = {col: str(col).lower().strip() for col in df.columns}

    mapping = {
        'name_col': None,
        'ic_col': None,
        'bank_col': None,
        'account_col': None,
        'position_col': None,
        'days_col': None,
        'date_col': None,
        'wage_col': None,
        'payment_col': None,
        'ot_col': None,
        'allowance_col': None,
        'transport_col': None,
        'claim_col': None,
        'total_col': None,
        'date_pattern_cols': []
    }

    for col, col_lower in cols.items():
        # Name column
        if 'name' in col_lower and 'bank' not in col_lower and not mapping['name_col']:
            mapping['name_col'] = col

        # IC column
        elif 'ic' in col_lower and not mapping['ic_col']:
            mapping['ic_col'] = col

        # Bank name
        elif ('bank name' in col_lower or col_lower == 'bank') and not mapping['bank_col']:
            mapping['bank_col'] = col

        # Account number
        elif 'account' in col_lower and not mapping['account_col']:
            mapping['account_col'] = col

        # Position
        elif col_lower in ['position', 'role'] and not mapping['position_col']:
            mapping['position_col'] = col

        # Days column - CRITICAL FIX
        elif col_lower in ['day', 'days'] and not mapping['days_col']:
            mapping['days_col'] = col

        # Date column
        elif col_lower == 'date' and not mapping['date_col']:
            mapping['date_col'] = col

        # Wages column
        elif col_lower in ['wages', 'wage', 'salary'] and not mapping['wage_col']:
            mapping['wage_col'] = col

        # Payment column (separate from wages)
        elif col_lower in ['payment'] and not mapping['payment_col']:
            mapping['payment_col'] = col

        # OT column
        elif col_lower in ['ot', 'overtime'] and not mapping['ot_col']:
            mapping['ot_col'] = col

        # Transport/Allowance
        elif col_lower in ['transport', 'transportation'] and not mapping['transport_col']:
            mapping['transport_col'] = col
        elif 'allowance' in col_lower and not mapping['allowance_col']:
            mapping['allowance_col'] = col

        # Claim column
        elif col_lower in ['claim', 'claims'] and not mapping['claim_col']:
            mapping['claim_col'] = col

        # Total column - USE THIS VALUE
        elif col_lower in ['total', 'total wages', 'total payment'] and not mapping['total_col']:
            mapping['total_col'] = col

        # Date pattern columns (for roster)
        elif re.match(r'\\d{4}-\\d{2}-\\d{2}', str(col)):
            mapping['date_pattern_cols'].append(col)

    return mapping


def is_continuation_row(row, prev_row, col_map):
    """
    Check if current row is a continuation of previous candidate.
    Continuation row = same candidate, additional data rows.
    """
    if prev_row is None:
        return False

    ic_col = col_map['ic_col']
    name_col = col_map['name_col']

    if not ic_col:
        return False

    # Current row IC is empty or same as previous
    curr_ic = clean_ic_number(row.get(ic_col))
    prev_ic = clean_ic_number(prev_row.get(ic_col))

    # If current IC is empty, check if name is also empty (strong signal of continuation)
    if not curr_ic:
        curr_name = clean_name(row.get(name_col)) if name_col else None
        # If name is also empty, likely continuation
        if not curr_name:
            return True

    # If IC matches previous, it's a continuation
    if curr_ic and prev_ic and curr_ic == prev_ic:
        return True

    return False


def process_excel_fixed(excel_path):
    """
    Process Excel file with FIXED continuation row handling.
    """
    print(f"\\n📂 Processing: {Path(excel_path).name}")

    month = extract_month_from_path(excel_path)
    all_data = []

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        print(f"   Found {len(sheet_names)} sheet(s)")

        for sheet_name in sheet_names:
            try:
                # Get project metadata
                metadata = extract_project_metadata(excel_path, sheet_name)

                # Read sheet
                df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)

                # Find header rows
                header_rows = []
                for idx, row in df.iterrows():
                    row_str = ' '.join([str(val) for val in row if pd.notna(val)]).lower()
                    if 'name' in row_str and 'ic' in row_str:
                        header_rows.append(idx)

                if not header_rows:
                    continue

                # Process each section (FIXED: prevent overlap)
                for section_idx, header_row_idx in enumerate(header_rows):
                    # Calculate correct nrows to avoid reading into next section
                    if section_idx + 1 < len(header_rows):
                        next_header = header_rows[section_idx + 1]
                        nrows = max(1, next_header - header_row_idx - 1)
                    else:
                        nrows = 200  # Last section, read all remaining
                    
                    section_df = pd.read_excel(
                        excel_path,
                        sheet_name=sheet_name,
                        skiprows=header_row_idx,
                        nrows=nrows
                    )

                    section_df.columns = [str(col).strip() for col in section_df.columns]

                    # Identify columns with better mapping
                    col_map = identify_columns(section_df)

                    if not col_map['name_col'] or not col_map['ic_col']:
                        continue

                    # Extract candidates with FIXED aggregation
                    current_candidate = None
                    candidates = {}
                    prev_row = None

                    for idx, row in section_df.iterrows():
                        # Check if this is a new candidate or continuation
                        is_continuation = is_continuation_row(row, prev_row, col_map)

                        if not is_continuation:
                            # New candidate row
                            if pd.notna(row.get(col_map['name_col'])) and pd.notna(row.get(col_map['ic_col'])):
                                if str(row[col_map['name_col']]).strip().lower() in ['no', 'name', '']:
                                    prev_row = row
                                    continue

                                ic_number = clean_ic_number(row[col_map['ic_col']])
                                if not ic_number or len(ic_number) < 6:
                                    prev_row = row
                                    continue

                                full_name = clean_name(row[col_map['name_col']])
                                if not full_name:
                                    prev_row = row
                                    continue

                                key = f"{ic_number}_{sheet_name}_{idx}"
                                current_candidate = key

                                candidates[key] = {
                                    'month': month,
                                    'source_file': Path(excel_path).name,
                                    'source_sheet': sheet_name,
                                    'project_name': sheet_name,
                                    'full_name': full_name,
                                    'alternate_name': extract_alternate_name(row[col_map['name_col']]),
                                    'ic_number': ic_number,
                                    'bank_name': clean_bank_name(row.get(col_map['bank_col'])) if col_map['bank_col'] else None,
                                    'account_number': clean_account(row.get(col_map['account_col'])) if col_map['account_col'] else None,
                                    'position': str(row.get(col_map['position_col'])).strip() if col_map['position_col'] and pd.notna(row.get(col_map['position_col'])) else None,
                                    'days_worked': 0,
                                    'total_wages': 0.0,
                                    'total_ot': 0.0,
                                    'total_allowance': 0.0,
                                    'total_claim': 0.0,
                                    'total_payment': 0.0,
                                    'work_dates': [],
                                    'project_date_range': metadata['project_date_range'],
                                    'payment_due_date': metadata['payment_due_date'],
                                    'location': metadata['location'],
                                    'time_schedule': metadata['time_schedule'],
                                    'roster_info': [],
                                    'notes': [],
                                    'project_notes': [],
                                    'payment_components': []  # Track for debugging
                                }

                        # Aggregate data from this row (whether new or continuation)
                        if current_candidate and current_candidate in candidates:
                            # Days
                            if col_map['days_col'] and pd.notna(row.get(col_map['days_col'])):
                                days_val = safe_float(row[col_map['days_col']])
                                candidates[current_candidate]['days_worked'] += days_val

                            # Wages
                            if col_map['wage_col'] and pd.notna(row.get(col_map['wage_col'])):
                                wage_val = safe_float(row[col_map['wage_col']])
                                candidates[current_candidate]['total_wages'] += wage_val
                                candidates[current_candidate]['payment_components'].append(f"Wage: {wage_val}")

                            # Payment (if separate from wages)
                            if col_map['payment_col'] and pd.notna(row.get(col_map['payment_col'])):
                                payment_val = safe_float(row[col_map['payment_col']])
                                candidates[current_candidate]['total_wages'] += payment_val
                                candidates[current_candidate]['payment_components'].append(f"Payment: {payment_val}")

                            # OT
                            if col_map['ot_col'] and pd.notna(row.get(col_map['ot_col'])):
                                ot_val = safe_float(row[col_map['ot_col']])
                                candidates[current_candidate]['total_ot'] += ot_val
                                candidates[current_candidate]['payment_components'].append(f"OT: {ot_val}")

                            # Transport/Allowance
                            if col_map['transport_col'] and pd.notna(row.get(col_map['transport_col'])):
                                transport_val = safe_float(row[col_map['transport_col']])
                                candidates[current_candidate]['total_allowance'] += transport_val
                                candidates[current_candidate]['payment_components'].append(f"Transport: {transport_val}")

                            if col_map['allowance_col'] and pd.notna(row.get(col_map['allowance_col'])):
                                allowance_val = safe_float(row[col_map['allowance_col']])
                                candidates[current_candidate]['total_allowance'] += allowance_val
                                candidates[current_candidate]['payment_components'].append(f"Allowance: {allowance_val}")

                            # Claim
                            if col_map['claim_col'] and pd.notna(row.get(col_map['claim_col'])):
                                claim_val = safe_float(row[col_map['claim_col']])
                                candidates[current_candidate]['total_claim'] += claim_val
                                candidates[current_candidate]['payment_components'].append(f"Claim: {claim_val}")

                            # Total - USE THIS if available, it's authoritative
                            if col_map['total_col'] and pd.notna(row.get(col_map['total_col'])):
                                total_val = safe_float(row[col_map['total_col']])
                                # Use the FIRST non-zero total we see
                                if total_val > 0 and candidates[current_candidate]['total_payment'] == 0:
                                    candidates[current_candidate]['total_payment'] = total_val
                                    candidates[current_candidate]['payment_components'].append(f"Total (from sheet): {total_val}")

                            # Date column
                            if col_map['date_col'] and pd.notna(row.get(col_map['date_col'])):
                                date_str = str(row[col_map['date_col']])
                                if date_str not in candidates[current_candidate]['work_dates']:
                                    candidates[current_candidate]['work_dates'].append(date_str)

                            # Roster info from date pattern columns
                            for date_col in col_map['date_pattern_cols']:
                                if pd.notna(row.get(date_col)):
                                    roster_val = str(row[date_col]).strip()
                                    if roster_val and len(roster_val) < 50:
                                        info = f"{date_col}: {roster_val}"
                                        if info not in candidates[current_candidate]['roster_info']:
                                            candidates[current_candidate]['roster_info'].append(info)

                        prev_row = row

                    # Post-process candidates
                    for key, candidate in candidates.items():
                        # Calculate total from components if not set
                        component_sum = (
                            candidate['total_wages'] +
                            candidate['total_ot'] +
                            candidate['total_allowance'] +
                            candidate['total_claim']
                        )

                        # If we didn't get a total from the "Total" column, use component sum
                        if candidate['total_payment'] == 0 and component_sum > 0:
                            candidate['total_payment'] = component_sum

                        # If component sum is larger (shouldn't happen, but just in case)
                        if component_sum > candidate['total_payment']:
                            candidate['total_payment'] = component_sum

                        # Limit roster info
                        candidate['roster_info'] = candidate['roster_info'][:5]

                        # Remove debugging field
                        candidate.pop('payment_components', None)

                        all_data.append(candidate)

            except Exception as e:
                print(f"   ⚠️  Error in sheet '{sheet_name}': {e}")
                continue

        print(f"   ✓ Extracted {len(all_data)} record(s)")
        return all_data

    except Exception as e:
        print(f"   ✗ Error: {e}")
        return []


def main():
    """Main function."""
    print("="*120)
    print(" "*30 + "FIXED EXTRACTION V3.2 - CONTINUATION ROW FIX")
    print("="*120)

    source_dir = "/Users/baito.kevin/Downloads/PROMOTER PAYMENT & CLAIMS 2025/Baito Promoter 2025/"

    months_ordered = [
        ('jan', 'Baito Jan Payment Details 2025.xlsx'),
        ('feb', 'Baito Feb Payment Details 2025.xlsx'),
        ('march', 'Baito March Payment Details 2025.xlsx'),
        ('april', 'Baito April Payment Details 2025.xlsx'),
        ('may', 'Baito May Payment Details 2025.xlsx'),
        ('june', 'Baito June Payment Details 2025.xlsx'),
        ('july', 'Baito July Payment Details 2025.xlsx'),
        ('aug', 'Baito Aug Payment Details 2025.xlsx'),
        ('sep', 'Baito Sep Payment Details 2025.xlsx')
    ]

    all_data = []

    for month_name, filename in months_ordered:
        excel_path = Path(source_dir) / filename

        if not excel_path.exists():
            print(f"\\n⊘ {month_name.capitalize()}: File not found")
            continue

        data = process_excel_fixed(str(excel_path))
        all_data.extend(data)

    print(f"\\n{'='*120}")
    print(f"Total records extracted: {len(all_data)}")
    print(f"{'='*120}")

    # Create DataFrame
    df = pd.DataFrame(all_data)

    # Format list columns
    df['work_dates'] = df['work_dates'].apply(lambda x: ', '.join(x) if x else '')
    df['roster_info'] = df['roster_info'].apply(lambda x: '; '.join(x) if x else '')
    df['notes'] = df['notes'].apply(lambda x: '; '.join(x) if x else '')
    df['project_notes'] = df['project_notes'].apply(lambda x: '; '.join(x) if x else '')

    # Month ordering for sorting
    month_order = {m.capitalize(): i for i, (m, _) in enumerate(months_ordered)}
    df['month_order'] = df['month'].map(month_order)

    # Sort by month (Jan to Dec), then file, then sheet
    df = df.sort_values(['month_order', 'source_file', 'source_sheet', 'full_name'])
    df = df.drop('month_order', axis=1)

    # Column order (ALL FIELDS)
    column_order = [
        'month',
        'source_file',
        'source_sheet',
        'project_name',
        'full_name',
        'alternate_name',
        'ic_number',
        'bank_name',
        'account_number',
        'position',
        'days_worked',
        'total_wages',
        'total_ot',
        'total_allowance',
        'total_claim',
        'total_payment',
        'work_dates',
        'project_date_range',
        'payment_due_date',
        'location',
        'time_schedule',
        'roster_info',
        'notes',
        'project_notes'
    ]

    df = df[column_order]

    # Fix account numbers as strings
    df['account_number'] = df['account_number'].astype('object')

    # Save
    output_file = 'baito_2025_FIXED_v3.2.xlsx'
    print(f"\\n💾 Saving to: {output_file}")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='All Candidates', index=False)

        # Format account_number column as text
        worksheet = writer.sheets['All Candidates']
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=9)  # Column I (account_number)
            if cell.value:
                cell.number_format = '@'
                cell.value = str(cell.value)

        # Monthly summary
        monthly = df.groupby('month').agg({
            'project_name': 'nunique',
            'full_name': 'count',
            'total_payment': 'sum'
        }).rename(columns={
            'project_name': 'Projects',
            'full_name': 'Records',
            'total_payment': 'Total Payment (RM)'
        })
        monthly.to_excel(writer, sheet_name='Monthly Summary')

    print(f"\\n✅ COMPLETE!")
    print(f"   File: {output_file}")
    print(f"   Records: {len(df):,}")
    print(f"   Columns: {len(df.columns)} (ALL metadata included)")
    print(f"   Sorted: Jan → Sep")

    # Verification for Chan Chiu Ling
    print(f"\\n🔍 VERIFICATION - Chan Chiu Ling (from Blackmores):")
    chan_records = df[(df['full_name'] == 'Chan Chiu Ling') & (df['source_sheet'] == 'Blackmores')]
    if len(chan_records) > 0:
        for idx, record in chan_records.iterrows():
            print(f"   Days: {record['days_worked']}")
            print(f"   Wages: {record['total_wages']}")
            print(f"   Total: {record['total_payment']}")
            print(f"   Expected: days=5, total=650")
    else:
        print("   ⚠️ Chan Chiu Ling not found in extraction")
    print()


if __name__ == '__main__':
    main()
