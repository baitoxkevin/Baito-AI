#!/usr/bin/env python3
"""
Automatically generate screenshots from Excel files for Vision AI processing
Converts each sheet to a high-quality PNG image

Usage:
    python excel_to_screenshots.py --input "excel_imports/*.xlsx" --output screenshots/

Requirements:
    pip install openpyxl pillow pandas xlsxwriter
"""

import argparse
import glob
import os
import sys
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from PIL import Image, ImageDraw, ImageFont
except ImportError as e:
    print(f"❌ Missing dependency: {e}")
    print("\n📦 Install required packages:")
    print("   pip install openpyxl pillow pandas xlsxwriter")
    sys.exit(1)


class ExcelToScreenshot:
    """Convert Excel sheets to screenshot images"""

    def __init__(self, output_dir: str, dpi: int = 150):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.dpi = dpi
        self.stats = {
            'files_processed': 0,
            'sheets_processed': 0,
            'images_created': 0,
            'errors': []
        }

    def process_excel_file(self, excel_path: str) -> list:
        """Process all sheets in an Excel file"""
        print(f"\n📂 Processing: {Path(excel_path).name}")

        try:
            wb = load_workbook(excel_path, data_only=True)
            images_created = []

            for sheet_name in wb.sheetnames:
                try:
                    print(f"   📄 Sheet: {sheet_name}")
                    image_path = self.render_sheet_to_image(wb, sheet_name, excel_path)

                    if image_path:
                        images_created.append(image_path)
                        self.stats['sheets_processed'] += 1
                        print(f"      ✅ Saved: {Path(image_path).name}")

                except Exception as e:
                    error_msg = f"Sheet '{sheet_name}' in {Path(excel_path).name}: {str(e)}"
                    self.stats['errors'].append(error_msg)
                    print(f"      ❌ Error: {str(e)}")

            self.stats['files_processed'] += 1
            return images_created

        except Exception as e:
            error_msg = f"File {Path(excel_path).name}: {str(e)}"
            self.stats['errors'].append(error_msg)
            print(f"   ❌ Failed to open file: {str(e)}")
            return []

    def render_sheet_to_image(self, workbook, sheet_name: str, source_file: str) -> str:
        """Render a single sheet as an image"""
        ws = workbook[sheet_name]

        # Calculate sheet dimensions
        max_row = ws.max_row
        max_col = ws.max_column

        if max_row == 0 or max_col == 0:
            print(f"      ⚠️  Empty sheet, skipping")
            return None

        # Create DataFrame from sheet
        data = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            data.append([cell.value for cell in row])

        df = pd.DataFrame(data)

        # Estimate image size (adjust based on content)
        cell_width = 100
        cell_height = 25
        img_width = min(max_col * cell_width, 3000)  # Max 3000px width
        img_height = min(max_row * cell_height, 5000)  # Max 5000px height

        # Create image
        img = Image.new('RGB', (img_width, img_height), color='white')
        draw = ImageDraw.Draw(img)

        # Try to load a monospace font, fallback to default
        try:
            font = ImageFont.truetype("/System/Library/Fonts/Monaco.ttf", 12)
            font_bold = ImageFont.truetype("/System/Library/Fonts/Monaco.ttf", 14)
        except:
            try:
                font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf", 12)
                font_bold = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSansMono-Bold.ttf", 14)
            except:
                font = ImageFont.load_default()
                font_bold = ImageFont.load_default()

        # Draw cells
        for row_idx, row in enumerate(data):
            for col_idx, value in enumerate(row):
                x = col_idx * cell_width
                y = row_idx * cell_height

                # Cell border
                draw.rectangle(
                    [x, y, x + cell_width - 1, y + cell_height - 1],
                    outline='#CCCCCC',
                    width=1
                )

                # Cell value
                if value is not None:
                    text = str(value)
                    # Truncate long text
                    if len(text) > 15:
                        text = text[:12] + '...'

                    # Use bold for first row (likely headers)
                    cell_font = font_bold if row_idx == 0 else font

                    # Background for header row
                    if row_idx == 0:
                        draw.rectangle(
                            [x + 1, y + 1, x + cell_width - 1, y + cell_height - 1],
                            fill='#E8E8E8'
                        )

                    # Merged cell indicator (approximate)
                    cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
                    if hasattr(cell, 'fill') and cell.fill.start_color.rgb and cell.fill.start_color.rgb != '00000000':
                        draw.rectangle(
                            [x + 1, y + 1, x + cell_width - 1, y + cell_height - 1],
                            fill='#FFF9E6'
                        )

                    # Draw text
                    draw.text(
                        (x + 5, y + 5),
                        text,
                        fill='black',
                        font=cell_font
                    )

        # Generate filename
        source_basename = Path(source_file).stem
        safe_sheet_name = sheet_name.replace('/', '_').replace(' ', '_')
        output_filename = f"{source_basename}_{safe_sheet_name}.png"
        output_path = self.output_dir / output_filename

        # Save image
        img.save(output_path, dpi=(self.dpi, self.dpi))
        self.stats['images_created'] += 1

        return str(output_path)

    def print_summary(self):
        """Print processing summary"""
        print("\n" + "═" * 60)
        print("📊 SCREENSHOT GENERATION SUMMARY")
        print("═" * 60)
        print(f"\n   Files Processed: {self.stats['files_processed']}")
        print(f"   Sheets Processed: {self.stats['sheets_processed']}")
        print(f"   Images Created: {self.stats['images_created']}")
        print(f"   Output Directory: {self.output_dir}")

        if self.stats['errors']:
            print(f"\n   ❌ Errors: {len(self.stats['errors'])}")
            for error in self.stats['errors'][:5]:
                print(f"      - {error}")
            if len(self.stats['errors']) > 5:
                print(f"      ... and {len(self.stats['errors']) - 5} more")

        print("\n" + "═" * 60)
        print("\n🎯 Next Steps:")
        print(f"   1. Review screenshots in: {self.output_dir}")
        print("   2. Run Vision AI extraction:")
        print(f"      node n8n-setup/batch-process-screenshots.js {self.output_dir}")
        print("")


def main():
    parser = argparse.ArgumentParser(
        description='Convert Excel files to screenshots for Vision AI processing'
    )
    parser.add_argument(
        '--input',
        required=True,
        help='Excel files to process (supports glob patterns like "*.xlsx")'
    )
    parser.add_argument(
        '--output',
        default='excel_screenshots',
        help='Output directory for screenshots (default: excel_screenshots)'
    )
    parser.add_argument(
        '--dpi',
        type=int,
        default=150,
        help='Image DPI (default: 150)'
    )

    args = parser.parse_args()

    # Find Excel files
    excel_files = glob.glob(args.input)

    if not excel_files:
        print(f"❌ No Excel files found matching: {args.input}")
        sys.exit(1)

    print("🚀 Excel to Screenshot Converter")
    print("━" * 60)
    print(f"📁 Input pattern: {args.input}")
    print(f"📂 Output directory: {args.output}")
    print(f"📸 Files found: {len(excel_files)}")
    print(f"🎨 DPI: {args.dpi}")
    print("━" * 60)

    # Process files
    converter = ExcelToScreenshot(args.output, args.dpi)

    for excel_file in excel_files:
        converter.process_excel_file(excel_file)

    # Summary
    converter.print_summary()


if __name__ == '__main__':
    main()
