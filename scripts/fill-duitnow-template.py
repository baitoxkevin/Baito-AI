#!/usr/bin/env python3
"""
DuitNow Template Filler
Fills the Excel template with payment data while preserving ALL formatting
"""

import sys
import json
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

def fill_duitnow_template(template_path, output_path, payment_data):
    """
    Fill DuitNow template with payment data

    Args:
        template_path: Path to the template Excel file
        output_path: Path where to save the filled Excel
        payment_data: Dict with 'paymentDate' and 'payments' array
    """
    # Load template (this preserves ALL formatting automatically)
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Fill payment date in A2 (or wherever it should go based on template)
    payment_date = payment_data.get('paymentDate', datetime.now().strftime('%d/%m/%Y'))

    # Check if B1 exists for payment date
    if ws['B1'].value is None:
        ws['A2'] = payment_date
    else:
        ws['B1'] = payment_date

    # Start filling data from row 4 (after headers and specs in rows 1-3)
    start_row = 4

    for idx, payment in enumerate(payment_data['payments']):
        row_num = start_row + idx

        # Column A: Payment Type
        ws[f'A{row_num}'] = payment.get('paymentType', 'DTN')

        # Column B: ID Type code
        ws[f'B{row_num}'] = payment.get('idType', 'NI')

        # Column C: BIC (empty for DuitNow)
        ws[f'C{row_num}'] = payment.get('bic', '')

        # Column D: Recipient's DuitNow ID
        ws[f'D{row_num}'] = payment.get('recipientId', '')

        # Column E: Payment Amount
        ws[f'E{row_num}'] = payment.get('amount', '')

        # Column F: Recipient Reference
        ws[f'F{row_num}'] = payment.get('reference', '')

        # Column G: Other Payment Details
        ws[f'G{row_num}'] = payment.get('paymentDetails', '')

        # Column H: Email 1
        ws[f'H{row_num}'] = payment.get('email1', '')

        # Column I: Email 2
        ws[f'I{row_num}'] = payment.get('email2', '')

        # Column J: Mobile 1
        ws[f'J{row_num}'] = payment.get('mobile1', '')

        # Column K: Mobile 2
        ws[f'K{row_num}'] = payment.get('mobile2', '')

    # Save the filled template
    wb.save(output_path)
    print(f"✓ Successfully created: {output_path}")

if __name__ == '__main__':
    if len(sys.argv) < 4:
        print("Usage: python3 fill-duitnow-template.py <template_path> <output_path> <json_data>")
        sys.exit(1)

    template_path = sys.argv[1]
    output_path = sys.argv[2]
    json_data = sys.argv[3]

    payment_data = json.loads(json_data)
    fill_duitnow_template(template_path, output_path, payment_data)
