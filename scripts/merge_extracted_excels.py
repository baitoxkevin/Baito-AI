#!/usr/bin/env python3
"""
Merge all Vision AI extracted Excel files into a single master file

Usage:
    python merge_extracted_excels.py --input "baito_extracted_*.xlsx" --output master.xlsx
"""

import argparse
import glob
import sys
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError as e:
    print(f"❌ Missing dependency: {e}")
    print("\n📦 Install required packages:")
    print("   pip install openpyxl pandas")
    sys.exit(1)


def merge_extracted_files(input_pattern: str, output_file: str):
    """Merge multiple extracted Excel files into one master file"""

    print("🔗 Merging Extracted Excel Files\n")
    print("━" * 60)

    # Find all files
    files = sorted(glob.glob(input_pattern))

    if not files:
        print(f"❌ No files found matching: {input_pattern}")
        sys.exit(1)

    print(f"📁 Files found: {len(files)}")
    for f in files:
        print(f"   - {Path(f).name}")

    print("\n⏳ Reading and merging...\n")

    # Read and merge all files
    all_data = []
    stats = {
        'total_files': len(files),
        'total_records': 0,
        'total_payment': 0,
        'files_processed': 0,
        'errors': []
    }

    for file_path in files:
        try:
            print(f"   📖 Reading: {Path(file_path).name}")

            # Read Excel file
            df = pd.read_excel(file_path)

            # Add source file column
            df['Source File'] = Path(file_path).stem

            all_data.append(df)

            # Update stats
            stats['files_processed'] += 1
            stats['total_records'] += len(df)

            # Calculate payment sum if column exists
            if 'Total Payment' in df.columns:
                payment_sum = df['Total Payment'].sum()
                stats['total_payment'] += payment_sum
                print(f"      Records: {len(df)}, Payment: RM {payment_sum:,.2f}")
            else:
                print(f"      Records: {len(df)}")

        except Exception as e:
            error_msg = f"{Path(file_path).name}: {str(e)}"
            stats['errors'].append(error_msg)
            print(f"      ❌ Error: {str(e)}")

    if not all_data:
        print("\n❌ No data to merge!")
        sys.exit(1)

    # Combine all DataFrames
    print("\n🔗 Combining data...")
    merged_df = pd.concat(all_data, ignore_index=True)

    # Sort by Payment Date (if exists) or Source File
    if 'Payment Date' in merged_df.columns:
        merged_df['Payment Date'] = pd.to_datetime(merged_df['Payment Date'], errors='coerce')
        merged_df = merged_df.sort_values(['Payment Date', 'Name'], na_position='last')
    elif 'Name' in merged_df.columns:
        merged_df = merged_df.sort_values('Name')

    # Save to Excel with formatting
    print(f"\n💾 Saving to: {output_file}")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write data
        merged_df.to_excel(writer, sheet_name='Merged Data', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Merged Data']

        # Format header row
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        worksheet.freeze_panes = 'A2'

        # Create summary sheet
        summary_df = pd.DataFrame([
            ['Total Files Merged', stats['total_files']],
            ['Total Records', stats['total_records']],
            ['Total Payment', f"RM {stats['total_payment']:,.2f}"],
            ['Files Processed Successfully', stats['files_processed']],
            ['Errors', len(stats['errors'])],
            [''],
            ['Source Files:'],
        ] + [[Path(f).name, ''] for f in files])

        summary_df.to_excel(writer, sheet_name='Summary', header=False, index=False)

    # Final summary
    print("\n" + "═" * 60)
    print("✅ MERGE COMPLETE")
    print("═" * 60)
    print(f"\n   Output File: {output_file}")
    print(f"   Total Records: {stats['total_records']:,}")
    print(f"   Total Payment: RM {stats['total_payment']:,.2f}")
    print(f"   Files Merged: {stats['files_processed']}/{stats['total_files']}")

    if stats['errors']:
        print(f"\n   ❌ Errors: {len(stats['errors'])}")
        for error in stats['errors']:
            print(f"      - {error}")

    # Data quality checks
    print("\n🔍 Data Quality Checks:")

    if 'IC Number' in merged_df.columns:
        unique_ics = merged_df['IC Number'].nunique()
        print(f"   Unique IC Numbers: {unique_ics}")

        if unique_ics < stats['total_records']:
            duplicates = stats['total_records'] - unique_ics
            print(f"   ⚠️  Duplicate IC Numbers: {duplicates}")

    if 'Confidence' in merged_df.columns:
        confidence_counts = merged_df['Confidence'].value_counts()
        print(f"\n   Confidence Distribution:")
        for conf, count in confidence_counts.items():
            print(f"      {conf}: {count} ({count / len(merged_df) * 100:.1f}%)")

    if 'Issues' in merged_df.columns:
        records_with_issues = merged_df['Issues'].notna().sum()
        if records_with_issues > 0:
            print(f"\n   ⚠️  Records with Issues: {records_with_issues}")

    print("\n" + "═" * 60)
    print("\n🎯 Next Steps:")
    print("   1. Open the merged file and review data")
    print("   2. Check records with low confidence or issues")
    print("   3. Validate against source Excel files if needed")
    print("   4. Import to Supabase when ready")
    print("")


def main():
    parser = argparse.ArgumentParser(
        description='Merge Vision AI extracted Excel files'
    )
    parser.add_argument(
        '--input',
        required=True,
        help='Input files pattern (e.g., "baito_extracted_*.xlsx")'
    )
    parser.add_argument(
        '--output',
        default='baito_2025_VISION_MASTER.xlsx',
        help='Output file name (default: baito_2025_VISION_MASTER.xlsx)'
    )

    args = parser.parse_args()

    merge_extracted_files(args.input, args.output)


if __name__ == '__main__':
    main()
